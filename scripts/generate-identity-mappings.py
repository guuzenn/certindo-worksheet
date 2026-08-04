from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


ROOT = Path(__file__).resolve().parents[1]
CATALOG_PATH = ROOT / "storage" / "generated" / "instrument-form-catalog.json"
OUTPUT_PATH = ROOT / "prisma" / "generated-identity-mappings.json"
DEFAULT_WORKBOOK = "storage/templates/Lembar Kerja 095-163.xlsx"


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def anchor_coordinate(sheet: Any, row: int, column: int) -> str:
    cell = sheet.cell(row, column)
    if not isinstance(cell, MergedCell):
        return cell.coordinate
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return sheet.cell(merged.min_row, merged.min_col).coordinate
    return cell.coordinate


def right_target(sheet: Any, cell: Any) -> str | None:
    end_column = cell.column
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= cell.row <= merged.max_row and merged.min_col <= cell.column <= merged.max_col:
            end_column = merged.max_col
            break
    if end_column >= sheet.max_column:
        return None
    return anchor_coordinate(sheet, cell.row, end_column + 1)


def add(mapping: dict[str, list[str]], key: str, coordinate: str | None) -> None:
    if coordinate and coordinate not in mapping.setdefault(key, []):
        mapping[key].append(coordinate)


def environment_targets(sheet: Any, label_cell: Any, prefix: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for header_row in range(max(1, label_cell.row - 3), label_cell.row + 1):
        for column in range(label_cell.column + 1, min(sheet.max_column, label_cell.column + 8) + 1):
            header = normalize(sheet.cell(header_row, column).value)
            if header == "awal":
                result[f"environment.{prefix}Start"] = anchor_coordinate(sheet, label_cell.row, column)
            elif header == "tengah":
                result[f"environment.{prefix}Middle"] = anchor_coordinate(sheet, label_cell.row, column)
            elif header == "akhir":
                result[f"environment.{prefix}End"] = anchor_coordinate(sheet, label_cell.row, column)
    return result


def mapping_for_sheet(sheet: Any) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    certificate_rows = [
        cell.row
        for row in sheet.iter_rows()
        for cell in row
        if not isinstance(cell, MergedCell) and normalize(cell.value) in {"no sertifikat", "nomor sertifikat"}
    ]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in (None, "") or isinstance(cell, MergedCell):
                continue
            label = normalize(cell.value)
            key: str | None = None
            if label in {"no sertifikat", "nomor sertifikat"}:
                key = "certificateNumber"
            elif label in {"tanggal kalibrasi", "tgl kalibrasi"}:
                key = "calibrationDate"
            elif label in {"nama alat", "alat"}:
                key = "instrument.name"
            elif label in {"merk", "merek"}:
                key = "instrument.manufacturer"
            elif label in {"type model", "tipe model", "model type", "model tipe", "model"}:
                key = "instrument.model"
            elif label in {"no seri", "nomor seri", "serial number", "no serial"}:
                key = "instrument.serialNumber"
            elif label in {"no identitas", "nomor identitas", "identitas"}:
                key = "instrument.identityNumber"
            elif label in {"kapasitas min", "kapasitas minimum"}:
                key = "instrument.capacityMin"
            elif label in {"kapasitas max", "kapasitas maksimum"}:
                key = "instrument.capacityMax"
            elif label == "kapasitas":
                key = "instrument.capacity"
            elif label == "resolusi":
                key = "instrument.resolution"
            elif label in {"nama perusahaan", "perusahaan"}:
                key = "company.name"
            elif label in {"lokasi kalibrasi", "tempat kalibrasi"}:
                key = "calibrationLocation"

            inside_identity_block = any(start <= cell.row <= start + 15 for start in certificate_rows)
            if key and (key == "certificateNumber" or inside_identity_block):
                add(mapping, key, right_target(sheet, cell))

            if inside_identity_block and label in {"temperature ruangan", "temperature ruang", "temperatur ruangan", "temperatur ruang", "suhu"}:
                for env_key, coordinate in environment_targets(sheet, cell, "temperature").items():
                    add(mapping, env_key, coordinate)
            elif inside_identity_block and label in {"kelembaban", "kelembapan"}:
                for env_key, coordinate in environment_targets(sheet, cell, "humidity").items():
                    add(mapping, env_key, coordinate)
    return mapping


def main() -> None:
    forms = json.loads(CATALOG_PATH.read_text(encoding="utf-8"))
    workbooks: dict[str, Any] = {}
    output: dict[str, dict[str, list[str]]] = {}
    for form in forms:
        workbook_path = form.get("workbook") or DEFAULT_WORKBOOK
        if workbook_path not in workbooks:
            workbooks[workbook_path] = load_workbook(ROOT / workbook_path, read_only=False, data_only=False)
        workbook = workbooks[workbook_path]
        sheet_name = form["sheet"]
        if sheet_name not in workbook.sheetnames:
            continue
        output[form["code"]] = mapping_for_sheet(workbook[sheet_name])
    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    for workbook in workbooks.values():
        workbook.close()
    print(json.dumps({"forms": len(forms), "mapped": len(output), "output": str(OUTPUT_PATH)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
