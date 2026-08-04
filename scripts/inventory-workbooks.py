from __future__ import annotations

import json
import hashlib
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = ROOT / "storage" / "templates"
OUTPUT_PATH = ROOT / "storage" / "generated" / "workbook-inventory.json"
CODE_PATTERN = re.compile(r"CCI-KAL-FOM[- ]*(0XX|XXX|\d{2,3})", re.IGNORECASE)


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_codes(value: Any) -> set[str]:
    if value is None:
        return set()
    text = str(value)
    codes: set[str] = set()
    for match in CODE_PATTERN.finditer(text):
        code = match.group(1).upper()
        if code.isdigit():
            code = code.zfill(3)
        codes.add(code)
    return codes


def inventory_sheet(sheet: Any) -> dict[str, Any]:
    formula_count = 0
    non_empty_count = 0
    error_literals: list[dict[str, str]] = []
    samples: list[dict[str, Any]] = []
    detected_codes = extract_codes(sheet.title)
    torque_cells: list[dict[str, Any]] = []

    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            non_empty_count += 1
            detected_codes.update(extract_codes(value))
            if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                formula_count += 1
            if isinstance(value, str) and value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}:
                error_literals.append({"cell": cell.coordinate, "value": value})
            if len(samples) < 50:
                samples.append({"cell": cell.coordinate, "value": json_value(value)})

    if "152" in detected_codes or "TORQUE GAUGE" in sheet.title.upper():
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    torque_cells.append({
                        "cell": cell.coordinate,
                        "value": json_value(cell.value),
                        "dataType": cell.data_type,
                        "numberFormat": cell.number_format,
                    })

    hidden_rows = [index for index, dim in sheet.row_dimensions.items() if dim.hidden]
    hidden_columns = [index for index, dim in sheet.column_dimensions.items() if dim.hidden]
    merged_ranges = [str(item) for item in sheet.merged_cells.ranges]
    merged_regions = []
    for item in sheet.merged_cells.ranges:
        anchor = sheet.cell(item.min_row, item.min_col)
        merged_regions.append({
            "range": str(item),
            "anchor": anchor.coordinate,
            "value": json_value(anchor.value),
        })
    validations = getattr(sheet.data_validations, "dataValidation", [])

    return {
        "name": sheet.title,
        "state": sheet.sheet_state,
        "dimension": sheet.calculate_dimension(),
        "maxRow": sheet.max_row,
        "maxColumn": sheet.max_column,
        "nonEmptyCells": non_empty_count,
        "formulaCells": formula_count,
        "literalFormulaErrors": error_literals,
        "detectedCodes": sorted(detected_codes),
        "mergedRangeCount": len(merged_ranges),
        "mergedRanges": merged_ranges,
        "mergedRegions": merged_regions,
        "printArea": str(sheet.print_area) if sheet.print_area else None,
        "printTitleRows": sheet.print_title_rows,
        "printTitleColumns": sheet.print_title_cols,
        "orientation": sheet.page_setup.orientation,
        "paperSize": sheet.page_setup.paperSize,
        "fitToWidth": sheet.page_setup.fitToWidth,
        "fitToHeight": sheet.page_setup.fitToHeight,
        "freezePanes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "imageCount": len(getattr(sheet, "_images", [])),
        "chartCount": len(getattr(sheet, "_charts", [])),
        "dataValidationCount": len(validations),
        "hiddenRows": hidden_rows,
        "hiddenColumns": hidden_columns,
        "sampleCells": samples,
        "torqueGaugeCells": torque_cells,
    }


def main() -> None:
    workbook_paths = sorted(TEMPLATE_DIR.glob("*.xlsx"))
    docx_paths = sorted((TEMPLATE_DIR / "docx").glob("*.docx"))
    docx_by_code: dict[str, list[str]] = defaultdict(list)
    for path in docx_paths:
        codes = extract_codes(path.name)
        for code in codes:
            docx_by_code[code].append(path.name)

    workbooks: list[dict[str, Any]] = []
    workbook_code_counts: Counter[str] = Counter()
    for path in workbook_paths:
        before = {"size": path.stat().st_size, "modified": path.stat().st_mtime_ns, "sha256": sha256(path)}
        workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
        sheets = [inventory_sheet(sheet) for sheet in workbook.worksheets]
        for sheet in sheets:
            workbook_code_counts.update(sheet["detectedCodes"])
        defined_names = [
            {"name": name, "target": item.attr_text, "broken": item.attr_text == "#REF!"}
            for name, item in workbook.defined_names.items()
        ]
        after = {"size": path.stat().st_size, "modified": path.stat().st_mtime_ns, "sha256": sha256(path)}
        workbooks.append({
            "file": path.name,
            "path": str(path.relative_to(ROOT)).replace("\\", "/"),
            "sheetCount": len(sheets),
            "sheetNames": workbook.sheetnames,
            "definedNameCount": len(workbook.defined_names),
            "definedNames": defined_names,
            "brokenDefinedNameCount": sum(1 for item in defined_names if item["broken"]),
            "sourceUnchangedDuringRead": before == after,
            "sheets": sheets,
        })
        workbook.close()

    all_docx_codes = set(docx_by_code)
    all_workbook_codes = set(workbook_code_counts)
    output = {
        "generatedFrom": "read-only openpyxl inventory; source workbooks were never saved",
        "workbookCount": len(workbook_paths),
        "docxCount": len(docx_paths),
        "workbooks": workbooks,
        "docxByCode": dict(sorted(docx_by_code.items())),
        "duplicateDocxCodes": {code: names for code, names in sorted(docx_by_code.items()) if len(names) > 1},
        "codesDetectedInWorkbooks": sorted(all_workbook_codes),
        "docxCodesNotDetectedInWorkbooks": sorted(all_docx_codes - all_workbook_codes),
        "workbookCodesWithoutDocx": sorted(all_workbook_codes - all_docx_codes),
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "output": str(OUTPUT_PATH),
        "workbooks": [{"file": item["file"], "sheetCount": item["sheetCount"]} for item in workbooks],
        "docxCount": len(docx_paths),
        "codesDetected": len(all_workbook_codes),
        "sourceUnchanged": all(item["sourceUnchangedDuringRead"] for item in workbooks),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
